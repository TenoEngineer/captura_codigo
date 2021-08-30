import os
from openpyxl import load_workbook
import win32com.client as win32
import glob

class LoopPast:
    #Funções para o loop para buscar o arquivo final

    def __init__(self, path_past, name_past):
        self.path_past = path_past
        self.name_past = name_past

    def find_past(self):
        if self.name_past is None:
            pass
        elif '.' not in self.name_past:
            new_path=os.path.join(self.path_past, self.name_past)
            new_past=os.listdir(new_path)
            self.path_past = new_path
            self.name_past = new_past
            return self.path_past, self.name_past

    def find_doc(self):
        doc=None
        for x in self.name_past:
            if 'DOCUMENT' in x.upper():
                doc=x
            if doc is None:
                return True
            else:
                return LoopPast(self.path_past, doc).find_past()

    def find_rm(self):
        x = None
        for x in self.name_past:
            if x is None:
                return True
            elif 'MATERIA' in x.upper():
                return LoopPast(self.path_past, x).find_past()

    def find_excel(self):
        for cod in self.name_past:
            if '.xlsx' in self.name_past:
                if cod[0]=='~':  #bug de arquivos com ~ a frente
                    next
                else:
                    excel=load_workbook(os.path.join(self.path_past,cod)) #Seleciona o arquivo
                    excel.active
                    aba= excel.sheetnames
                    planilha = excel[aba[0]]  #Seleciona a aba
                    cod = planilha['I1'].value  #Copia valor

                if len(str(cod)) == 9 or len(str(cod)) == 10:
                    cod = planilha['L2'].value  #Copia valor

                if cod == '(51)-':
                    cod = planilha['L1'].value  #Copia valor

                #Testes para ver se o cod pegou certo
                if not cod:
                    next
                elif len(str(cod)) > 50:  #Caso haja alguma fórmula
                    cod = None
                elif len(str(cod)) > 28:
                    y=(cod.split()[-1]).split(sep='-')
                    if len(y)<2:
                        cod = None
                    else:
                        cod = y[0][:2]+ '_' + y[-2][-3:]   #Novo código
                else:
                    cod = None

            else:
                next

        if cod is None:
            return self
        else:
            return cod

    def xls_to_xlsx(self):
        o=win32.gencache.EnsureDispatch("Excel.Application")
        o.Visible = False
        input_dir = self.path_past
        output_dir = self.path_past
        files = glob.glob(input_dir + "/*.xls")
        files_xlsx = glob.glob(input_dir + '/*.xlsx')
        for filename in files:
            file = os.path.basename(filename)
            output = (output_dir + '\\' + file).replace('.xls', '.xlsx')
            output=output.replace({'[':'(', ']':')'})
            if output in files_xlsx:
                #print('Já existe arquivo xlsx, não precisa substituir \\n')
                next
            else:
                wb = o.Workbooks.Open(filename)
                wb.SaveAs(output, FileFormat = 51)
                wb.Close(True)
                #print('Arquivo salvo como xlsx com sucesso {file} \\n'.format(file=output))
                next
